from pathlib import Path
import os
import uuid
import datetime
from functools import wraps
import mimetypes
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
try:
    import pdfkit
    _PDFKIT_AVAILABLE = True
except Exception:
    _PDFKIT_AVAILABLE = False
try:
    from weasyprint import HTML as WeasyHTML
    _WEASY_AVAILABLE = True
except Exception:
    _WEASY_AVAILABLE = False
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import inspect, or_, text
from models import db, Bid, RequirementDocument, PreDocument, LoginLog

BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = BASE_DIR / 'app.db'
UPLOAD_FOLDER = BASE_DIR / 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'png', 'jpg', 'jpeg', 'txt'}
PHILGEPS_CATEGORIES = [
    'School Supplies',
    'Janitorial Supplies',
    'Office Supplies',
    'Furniture and Fixtures',
    'Food and Beverage',
    'IT Equipment',
    'Medical Supplies',
    'Construction Materials',
    'Laboratory Equipment',
    'Security Services',
    'Consultancy Services',
    'Printing Services',
    'Maintenance Services',
    'Transportation Services',
    'PPE and Safety Equipment',
    'Communication Equipment',
    'Electrical Supplies',
    'Stationery',
    'Packaging Materials',
    'Cleaning Products',
]

if not UPLOAD_FOLDER.exists():
    UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)


def create_app():
    app = Flask(__name__, static_folder='static', template_folder='templates')
    app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'dev')
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + str(DATABASE_PATH).replace('\\', '/')
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)
    app.config['ADMIN_USERNAME'] = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin')
    app.config['ADMIN_PASSWORD_HASH'] = generate_password_hash(admin_password)

    db.init_app(app)

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def save_attachments(files):
        saved_names = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                unique_name = f"{uuid.uuid4().hex}_{filename}"
                file_path = UPLOAD_FOLDER / unique_name
                file.save(file_path)
                saved_names.append(unique_name)
        return saved_names

    def parse_date_input(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return None

    def resolve_categories(existing_values):
        combined = list(PHILGEPS_CATEGORIES) + [value for value in existing_values if value]
        return sorted(dict.fromkeys(combined), key=lambda x: x.lower())

    def login_required(view):
        @wraps(view)
        def wrapped_view(**kwargs):
            if session.get('user'):
                return view(**kwargs)
            return redirect(url_for('login', next=request.path))
        return wrapped_view

    def form_data_from_bid(bid):
        return {
            'date_monitored': bid.date_monitored.strftime('%Y-%m-%d') if bid.date_monitored else '',
            'reference_number': bid.reference_number or '',
            'system_type': bid.system_type or 'Philgeps',
            'control_number': bid.control_number or '',
            'procuring_entity': bid.procuring_entity or '',
            'project_title': bid.project_title or '',
            'category': bid.category or '',
            'region': bid.region or '',
            'province': bid.province or '',
            'delivery_location': bid.delivery_location or '',
            'contact_person': bid.contact_person or '',
            'abc_value': bid.abc_value if bid.abc_value is not None else '',
            'procurement_mode': bid.procurement_mode or '',
            'date_created': bid.date_created.strftime('%Y-%m-%dT%H:%M') if bid.date_created else '',
            'closing_date': bid.closing_date.strftime('%Y-%m-%d') if bid.closing_date else '',
            'closing_time': bid.closing_time or '',
            'days_remaining': bid.days_remaining if bid.days_remaining is not None else '',
            'priority_level': bid.priority_level or '',
            'status': bid.status or '',
            'canvassing_status': bid.canvassing_status or '',
            'remarks': bid.remarks or '',
            'date_submitted': bid.date_submitted.strftime('%Y-%m-%d') if bid.date_submitted else '',
            'bid_amount': bid.bid_amount if bid.bid_amount is not None else '',
            'result': bid.result or '',
            'opportunity_score': bid.opportunity_score if bid.opportunity_score is not None else '',
            'abc_flag': bid.abc_flag or '',
            'recommendation': bid.recommendation or '',
            'attachments': bid.attachments or '',
        }

    def parse_float_input(value):
        if not value:
            return None
        try:
            normalized = str(value).replace('₱', '').replace(',', '').strip()
            return float(normalized)
        except ValueError:
            return None

    def parse_datetime_input(value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, '%Y-%m-%dT%H:%M')
        except ValueError:
            return None

    def ensure_db_schema():
        inspector = inspect(db.engine)
        if 'bids' in inspector.get_table_names():
            columns = [column['name'] for column in inspector.get_columns('bids')]
            if 'attachments' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN attachments VARCHAR(2048)'))
            if 'canvassing_status' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN canvassing_status VARCHAR(128)'))
            if 'date_added' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN date_added DATE'))
                    conn.execute(text("UPDATE bids SET date_added = DATE('now') WHERE date_added IS NULL"))
            if 'system_type' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text("ALTER TABLE bids ADD COLUMN system_type VARCHAR(64) DEFAULT 'Philgeps'"))
                    conn.execute(text("UPDATE bids SET system_type = 'Philgeps' WHERE system_type IS NULL"))
            if 'control_number' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN control_number VARCHAR(128)'))
            if 'procurement_mode' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN procurement_mode VARCHAR(128)'))
            if 'delivery_location' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN delivery_location VARCHAR(255)'))
            if 'contact_person' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN contact_person VARCHAR(128)'))
            if 'date_created' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN date_created DATETIME'))
        if 'requirements' in inspector.get_table_names():
            requirement_columns = [column['name'] for column in inspector.get_columns('requirements')]
        else:
            requirement_columns = []
        if 'requirements' not in inspector.get_table_names():
            with db.engine.begin() as conn:
                conn.execute(text('CREATE TABLE requirements (id INTEGER PRIMARY KEY, title VARCHAR(255) NOT NULL, description TEXT, filename VARCHAR(1024) NOT NULL UNIQUE, original_filename VARCHAR(1024) NOT NULL, upload_date DATETIME NOT NULL)'))

        if 'login_logs' not in inspector.get_table_names():
            with db.engine.begin() as conn:
                conn.execute(text('CREATE TABLE login_logs (id INTEGER PRIMARY KEY, username VARCHAR(128) NOT NULL, action VARCHAR(32) NOT NULL, ip_address VARCHAR(64), user_agent TEXT, timestamp DATETIME NOT NULL)'))

    with app.app_context():
        ensure_db_schema()

    @app.template_filter('format_date')
    def format_date(value):
        return value.strftime('%Y-%m-%d') if value else '-'

    @app.template_filter('format_datetime')
    def format_datetime(value):
        return value.strftime('%Y-%m-%d %I:%M %p') if value else '-'

    @app.template_filter('format_time')
    def format_time(value):
        if not value:
            return '-'
        try:
            parsed = datetime.datetime.strptime(value, '%H:%M')
            return parsed.strftime('%I:%M %p').lstrip('0')
        except ValueError:
            return value

    @app.template_filter('format_currency')
    def format_currency(value):
        if value is None:
            return '-'
        return f'₱{value:,.2f}'

    @app.route('/')
    def index():
        if session.get('user'):
            return redirect(url_for('dashboard'))
        return redirect(url_for('login'))

    @app.route('/dashboard')
    @login_required
    def dashboard():
        today = datetime.date.today()
        next_week = today + datetime.timedelta(days=7)

        total_bids = Bid.query.count()
        total_requirements = RequirementDocument.query.count()
        open_bids = Bid.query.filter(Bid.status.in_(['Pending', 'Submitted', 'Under Evaluation', 'Cancelled', 'For Review', 'Interested', 'For Canvas', 'New', 'Submitting'])).count()
        closing_soon = Bid.query.filter(Bid.closing_date.isnot(None), Bid.closing_date >= today, Bid.closing_date <= next_week).count()
        with_attachments = Bid.query.filter(Bid.attachments.isnot(None), Bid.attachments != '').count()

        status_counts = dict(db.session.query(Bid.status, db.func.count(Bid.id)).group_by(Bid.status).all())
        priority_counts = dict(db.session.query(Bid.priority_level, db.func.count(Bid.id)).group_by(Bid.priority_level).all())

        recent_bids = Bid.query.order_by(Bid.date_monitored.desc().nullslast(), Bid.closing_date.asc().nullslast()).limit(5).all()

        return render_template(
            'dashboard.html',
            total_bids=total_bids,
            total_requirements=total_requirements,
            open_bids=open_bids,
            closing_soon=closing_soon,
            with_attachments=with_attachments,
            status_counts=status_counts,
            priority_counts=priority_counts,
            recent_bids=recent_bids,
        )

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if session.get('user'):
            return redirect(url_for('dashboard'))

        error = None
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '').strip()
            remember = request.form.get('remember') == 'on'
            if username != app.config['ADMIN_USERNAME'] or not check_password_hash(app.config['ADMIN_PASSWORD_HASH'], password):
                error = 'Invalid username or password.'
                flash(error, 'danger')
            else:
                session['user'] = username
                session.permanent = remember
                log = LoginLog(
                    username=username,
                    action='login',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent'),
                )
                db.session.add(log)
                db.session.commit()
                flash('Logged in successfully.', 'success')
                next_page = request.args.get('next') or url_for('dashboard')
                return redirect(next_page)
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        username = session.pop('user', None)
        if username:
            log = LoginLog(
                username=username,
                action='logout',
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent'),
            )
            db.session.add(log)
            db.session.commit()
        flash('You have been logged out.', 'success')
        return redirect(url_for('login'))

    @app.route('/logbook')
    @login_required
    def logbook():
        page = request.args.get('page', 1, type=int)
        pagination = LoginLog.query.order_by(LoginLog.timestamp.desc()).paginate(
            page=page, per_page=25, error_out=False
        )
        return render_template('logbook.html', logs=pagination.items, pagination=pagination)

    @app.route('/uploads/<path:filename>')
    @login_required
    def uploaded_file(filename):
        return send_from_directory(UPLOAD_FOLDER, filename)

    @app.route('/bid-monitoring/new', methods=['GET', 'POST'])
    @login_required
    def bid_new():
        statuses = ['Pending', 'Submitted', 'Under Evaluation', 'Cancelled', 'For Review', 'Interested', 'For Canvas', 'New', 'Submitting', 'Awarded']
        priorities = ['Low', 'Medium', 'High', 'Urgent']
        regions = [row[0] for row in db.session.query(Bid.region).distinct().order_by(Bid.region).all() if row[0]]
        db_categories = [row[0] for row in db.session.query(Bid.category).distinct().order_by(Bid.category).all() if row[0]]
        categories = resolve_categories(db_categories)

        form_data = {}
        if request.method == 'POST':
            form_data = request.form.to_dict()
            reference_number = form_data.get('reference_number', '').strip()
            if not reference_number:
                flash('Reference number is required.', 'danger')
            else:
                closing_date = parse_date_input(form_data.get('closing_date'))
                status_value = form_data.get('status', '').strip() or None
                submitted_date = parse_date_input(form_data.get('date_submitted'))
                if status_value == 'Submitted' and not submitted_date:
                    submitted_date = datetime.date.today()
                elif status_value != 'Submitted':
                    submitted_date = None

                attachments = request.files.getlist('attachments')
                saved_files = save_attachments(attachments)
                days_remaining = None
                if closing_date:
                    days_remaining = (closing_date - datetime.date.today()).days

                bid = Bid(
                    date_monitored=parse_date_input(form_data.get('date_monitored')),
                    reference_number=reference_number,
                    system_type=form_data.get('system_type', '').strip() or 'Philgeps',
                    control_number=form_data.get('control_number', '').strip() or None,
                    procuring_entity=form_data.get('procuring_entity', '').strip() or None,
                    project_title=form_data.get('project_title', '').strip() or None,
                    category=form_data.get('category', '').strip() or None,
                    region=form_data.get('region', '').strip() or None,
                    province=form_data.get('province', '').strip() or None,
                    abc_value=parse_float_input(form_data.get('abc_value')),
                    delivery_location=form_data.get('delivery_location', '').strip() or None,
                    contact_person=form_data.get('contact_person', '').strip() or None,
                    procurement_mode=form_data.get('procurement_mode', '').strip() or None,
                    date_created=parse_datetime_input(form_data.get('date_created')),
                    closing_date=closing_date,
                    closing_time=form_data.get('closing_time', '').strip() or None,
                    days_remaining=days_remaining,
                    priority_level=form_data.get('priority_level', '').strip() or None,
                    status=status_value,
                    canvassing_status=form_data.get('canvassing_status', '').strip() or None,
                    remarks=form_data.get('remarks', '').strip() or None,
                    date_submitted=submitted_date,
                    bid_amount=parse_float_input(form_data.get('bid_amount')),
                    result=form_data.get('result', '').strip() or None,
                    opportunity_score=int(form_data.get('opportunity_score')) if form_data.get('opportunity_score') else None,
                    abc_flag=form_data.get('abc_flag', '').strip() or None,
                    recommendation=form_data.get('recommendation', '').strip() or None,
                    attachments=';'.join(saved_files) if saved_files else None,
                )
                db.session.add(bid)
                db.session.commit()
                flash('New bid created successfully.', 'success')
                return redirect(url_for('bid_detail', bid_id=bid.id))

        return render_template(
            'bid_form.html',
            form_data=form_data,
            statuses=statuses,
            priorities=priorities,
            regions=regions,
            categories=categories,
            form_action=url_for('bid_new'),
            edit_mode=False,
            bid=None,
        )

    @app.route('/bid-monitoring/<int:bid_id>/edit', methods=['GET', 'POST'])
    @login_required
    def bid_edit(bid_id):
        bid = Bid.query.get_or_404(bid_id)
        statuses = ['Pending', 'Submitted', 'Under Evaluation', 'Cancelled', 'For Review', 'Interested', 'For Canvas', 'New', 'Submitting', 'Awarded']
        priorities = ['Low', 'Medium', 'High', 'Urgent']
        regions = [row[0] for row in db.session.query(Bid.region).distinct().order_by(Bid.region).all() if row[0]]
        db_categories = [row[0] for row in db.session.query(Bid.category).distinct().order_by(Bid.category).all() if row[0]]
        categories = resolve_categories(db_categories)

        form_data = form_data_from_bid(bid)
        if request.method == 'POST':
            form_data = request.form.to_dict()
            reference_number = form_data.get('reference_number', '').strip()
            if not reference_number:
                flash('Reference number is required.', 'danger')
            else:
                existing_bid = Bid.query.filter(Bid.reference_number == reference_number, Bid.id != bid.id).first()
                if existing_bid:
                    flash('Another bid already uses that reference number.', 'danger')
                else:
                    closing_date = parse_date_input(form_data.get('closing_date'))
                    status_value = form_data.get('status', '').strip() or None
                    submitted_date = parse_date_input(form_data.get('date_submitted'))
                    if status_value == 'Submitted' and not submitted_date:
                        submitted_date = datetime.date.today()
                    elif status_value != 'Submitted':
                        submitted_date = None

                    attachments = request.files.getlist('attachments')
                    saved_files = save_attachments(attachments)
                    days_remaining = None
                    if closing_date:
                        days_remaining = (closing_date - datetime.date.today()).days

                    bid.date_monitored = parse_date_input(form_data.get('date_monitored'))
                    bid.reference_number = reference_number
                    bid.system_type = form_data.get('system_type', '').strip() or 'Philgeps'
                    bid.control_number = form_data.get('control_number', '').strip() or None
                    bid.procuring_entity = form_data.get('procuring_entity', '').strip() or None
                    bid.project_title = form_data.get('project_title', '').strip() or None
                    bid.category = form_data.get('category', '').strip() or None
                    bid.region = form_data.get('region', '').strip() or None
                    bid.province = form_data.get('province', '').strip() or None
                    bid.abc_value = parse_float_input(form_data.get('abc_value'))
                    bid.delivery_location = form_data.get('delivery_location', '').strip() or None
                    bid.contact_person = form_data.get('contact_person', '').strip() or None
                    bid.procurement_mode = form_data.get('procurement_mode', '').strip() or None
                    bid.date_created = parse_datetime_input(form_data.get('date_created'))
                    bid.closing_date = closing_date
                    bid.closing_time = form_data.get('closing_time', '').strip() or None
                    bid.days_remaining = days_remaining
                    bid.priority_level = form_data.get('priority_level', '').strip() or None
                    bid.status = status_value
                    bid.remarks = form_data.get('remarks', '').strip() or None
                    bid.date_submitted = submitted_date
                    bid.bid_amount = parse_float_input(form_data.get('bid_amount'))
                    bid.result = form_data.get('result', '').strip() or None
                    bid.opportunity_score = int(form_data.get('opportunity_score')) if form_data.get('opportunity_score') else None
                    bid.abc_flag = form_data.get('abc_flag', '').strip() or None
                    bid.recommendation = form_data.get('recommendation', '').strip() or None
                    bid.canvassing_status = form_data.get('canvassing_status', '').strip() or None
                    if saved_files:
                        existing = bid.attachments.split(';') if bid.attachments else []
                        bid.attachments = ';'.join(existing + saved_files)
                    db.session.commit()
                    flash('Bid updated successfully.', 'success')
                    return redirect(url_for('bid_detail', bid_id=bid.id))

        return render_template(
            'bid_form.html',
            form_data=form_data,
            statuses=statuses,
            priorities=priorities,
            regions=regions,
            categories=categories,
            form_action=url_for('bid_edit', bid_id=bid.id),
            edit_mode=True,
            bid=bid,
        )

    @app.route('/bid-monitoring/<int:bid_id>/delete', methods=['POST'])
    @login_required
    def bid_delete(bid_id):
        bid = Bid.query.get_or_404(bid_id)
        db.session.delete(bid)
        db.session.commit()
        flash('Bid deleted successfully.', 'success')
        return redirect(url_for('bid_monitoring'))

    @app.route('/requirements')
    @login_required
    def requirements_list():
        docs = RequirementDocument.query.order_by(RequirementDocument.upload_date.desc()).all()
        return render_template('requirements_list.html', docs=docs)

    @app.route('/requirements/new', methods=['GET', 'POST'])
    @login_required
    def requirement_new():
        if request.method == 'POST':
            title = request.form.get('title', '').strip()
            description = request.form.get('description', '').strip()
            file = request.files.get('file')
            if not title:
                flash('Title is required.', 'danger')
            elif not file or file.filename == '':
                flash('Requirement file is required.', 'danger')
            elif not allowed_file(file.filename):
                flash('File type not allowed.', 'danger')
            else:
                original_filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{original_filename}"
                file_path = UPLOAD_FOLDER / unique_filename
                file.save(file_path)
                doc = RequirementDocument(
                    title=title,
                    description=description or None,
                    filename=unique_filename,
                    original_filename=original_filename,
                )
                db.session.add(doc)
                db.session.commit()
                flash('Requirement saved successfully.', 'success')
                return redirect(url_for('requirements_list'))
        return render_template('requirement_form.html')

    @app.route('/requirements/<int:doc_id>')
    @login_required
    def requirement_detail(doc_id):
        doc = RequirementDocument.query.get_or_404(doc_id)
        return render_template('requirement_detail.html', doc=doc)

    @app.route('/requirements/<path:filename>/download')
    @login_required
    def requirement_download(filename):
        return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

    @app.route('/requirements/<path:filename>/view')
    @login_required
    def requirement_view(filename):
        if is_inline_viewable(filename):
            return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=False)
        doc = RequirementDocument.query.filter_by(filename=filename).first()
        if doc:
            return render_template('pre_document_preview.html', doc=doc)
        return redirect(url_for('requirements_list'))

    @app.route('/pre-documents')
    @login_required
    def pre_documents_list():
        docs = PreDocument.query.order_by(PreDocument.upload_date.desc()).all()
        return render_template('pre_documents_list.html', docs=docs)

    @app.route('/pre-documents/new', methods=['GET', 'POST'])
    @login_required
    def pre_document_new():
        if request.method == 'POST':
            title = request.form.get('title', '').strip()
            supplier_name = request.form.get('supplier_name', '').strip()
            description = request.form.get('description', '').strip()
            file = request.files.get('file')
            if not title:
                flash('Title is required.', 'danger')
            elif not file or file.filename == '':
                flash('Document file is required.', 'danger')
            elif not allowed_file(file.filename):
                flash('File type not allowed.', 'danger')
            else:
                original_filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{original_filename}"
                file_path = UPLOAD_FOLDER / unique_filename
                file.save(file_path)
                doc = PreDocument(
                    title=title,
                    supplier_name=supplier_name or None,
                    description=description or None,
                    filename=unique_filename,
                    original_filename=original_filename,
                )
                db.session.add(doc)
                db.session.commit()
                flash('Pre-document saved successfully.', 'success')
                return redirect(url_for('pre_documents_list'))
        return render_template('pre_document_form.html')

    @app.route('/pre-documents/<int:doc_id>')
    @login_required
    def pre_document_detail(doc_id):
        doc = PreDocument.query.get_or_404(doc_id)
        return render_template('pre_document_detail.html', doc=doc)

    @app.route('/pre-documents/<path:filename>/download')
    @login_required
    def pre_document_download(filename):
        return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

    def is_inline_viewable(filename):
        mime_type, _ = mimetypes.guess_type(filename)
        return mime_type in {
            'application/pdf',
            'image/png',
            'image/jpeg',
            'image/jpg',
            'text/plain'
        }

    @app.route('/pre-documents/<path:filename>/view')
    @login_required
    def pre_document_view(filename):
        if is_inline_viewable(filename):
            return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=False)
        doc = PreDocument.query.filter_by(filename=filename).first()
        if doc:
            return render_template('pre_document_preview.html', doc=doc)
        return redirect(url_for('pre_documents_list'))

    def build_bid_query_from_args(args):
        query = Bid.query
        search = args.get('search', '', type=str).strip()
        if search:
            like_term = f"%{search}%"
            query = query.filter(
                or_(
                    Bid.reference_number.ilike(like_term),
                    Bid.control_number.ilike(like_term),
                    Bid.project_title.ilike(like_term),
                    Bid.procuring_entity.ilike(like_term),
                )
            )

        date_added_filter = args.get('date_added', '', type=str).strip()
        if date_added_filter:
            try:
                dt = datetime.datetime.strptime(date_added_filter, '%Y-%m-%d').date()
                query = query.filter(Bid.date_added == dt)
            except ValueError:
                pass

        closing_date_filter = args.get('closing_date', '', type=str).strip()
        if closing_date_filter:
            try:
                cd = datetime.datetime.strptime(closing_date_filter, '%Y-%m-%d').date()
                query = query.filter(Bid.closing_date == cd)
            except ValueError:
                pass

        status_filter = args.get('status', '', type=str).strip()
        if status_filter:
            query = query.filter(Bid.status == status_filter)

        region_filter = args.get('region', '', type=str).strip()
        if region_filter:
            query = query.filter(Bid.region == region_filter)

        category_filter = args.get('category', '', type=str).strip()
        if category_filter:
            query = query.filter(Bid.category == category_filter)

        priority_filter = args.get('priority', '', type=str).strip()
        if priority_filter:
            query = query.filter(Bid.priority_level == priority_filter)

        system_type_filter = args.get('system_type', '', type=str).strip()
        if system_type_filter:
            query = query.filter(Bid.system_type == system_type_filter)

        return query

    @app.route('/reports/summary')
    @login_required
    def reports_summary():
        query = build_bid_query_from_args(request.args)
        bids = query.order_by(Bid.date_monitored.desc().nullslast(), Bid.closing_date.asc().nullslast()).all()
        # gather filter option lists
        statuses = [s[0] for s in db.session.query(Bid.status).distinct().order_by(Bid.status).all() if s[0]]
        regions = [r[0] for r in db.session.query(Bid.region).distinct().order_by(Bid.region).all() if r[0]]
        categories = [c[0] for c in db.session.query(Bid.category).distinct().order_by(Bid.category).all() if c[0]]
        system_types = [t[0] for t in db.session.query(Bid.system_type).distinct().order_by(Bid.system_type).all() if t[0]]
        priorities = [p[0] for p in db.session.query(Bid.priority_level).distinct().order_by(Bid.priority_level).all() if p[0]]
        canvassing_statuses = [c[0] for c in db.session.query(Bid.canvassing_status).distinct().order_by(Bid.canvassing_status).all() if c[0]]
        return render_template('summary_report.html', bids=bids, filters=request.args, generated_at=datetime.datetime.utcnow(), statuses=statuses, regions=regions, categories=categories, system_types=system_types, priorities=priorities, canvassing_statuses=canvassing_statuses)

    @app.route('/reports/summary/download_csv')
    @login_required
    def reports_summary_csv():
        query = build_bid_query_from_args(request.args)
        bids = query.order_by(Bid.date_monitored.desc().nullslast(), Bid.closing_date.asc().nullslast()).all()

        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Reference', 'Date Added', 'Project Title', 'Entity', 'System', 'Category', 'Region', 'ABC', 'Closing Date', 'Closing Time', 'Status', 'Canvassing', 'Priority'])
        for b in bids:
            cw.writerow([
                b.reference_number,
                b.date_added.strftime('%Y-%m-%d') if b.date_added else '',
                b.project_title or '',
                b.procuring_entity or '',
                b.system_type or '',
                b.category or '',
                b.region or '',
                f"{b.abc_value:.2f}" if b.abc_value is not None else '',
                b.closing_date.strftime('%Y-%m-%d') if b.closing_date else '',
                b.closing_time or '',
                b.status or '',
                b.canvassing_status or '',
                b.priority_level or '',
            ])
        output = si.getvalue()
        filename = f"summary_{datetime.date.today().isoformat()}.csv"
        return (output, 200, {
            'Content-Type': 'text/csv; charset=utf-8',
            'Content-Disposition': f'attachment; filename="{filename}"'
        })

    @app.route('/reports/summary/download_xlsx')
    @login_required
    def reports_summary_xlsx():
        query = build_bid_query_from_args(request.args)
        bids = query.order_by(Bid.date_monitored.desc().nullslast(), Bid.closing_date.asc().nullslast()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'

        header = ['Reference', 'Date Added', 'Project Title', 'Entity', 'System', 'Category', 'Region', 'ABC', 'Closing Date', 'Closing Time', 'Status', 'Canvassing', 'Priority']
        bold = Font(bold=True)
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        align = Alignment(horizontal='left', vertical='center')
        thin = Side(border_style='thin', color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
            cell.fill = fill
            cell.alignment = align
            cell.border = border

        for r, b in enumerate(bids, start=2):
            ws.cell(row=r, column=1, value=b.reference_number)
            ws.cell(row=r, column=2, value=b.date_added.strftime('%Y-%m-%d') if b.date_added else '')
            ws.cell(row=r, column=3, value=b.project_title or '')
            ws.cell(row=r, column=4, value=b.procuring_entity or '')
            ws.cell(row=r, column=5, value=b.system_type or '')
            ws.cell(row=r, column=6, value=b.category or '')
            ws.cell(row=r, column=7, value=b.region or '')
            abc_cell = ws.cell(row=r, column=8, value=b.abc_value if b.abc_value is not None else None)
            abc_cell.number_format = '#,##0.00'
            ws.cell(row=r, column=9, value=b.closing_date.strftime('%Y-%m-%d') if b.closing_date else '')
            ws.cell(row=r, column=10, value=b.closing_time or '')
            ws.cell(row=r, column=11, value=b.status or '')
            ws.cell(row=r, column=12, value=b.canvassing_status or '')
            ws.cell(row=r, column=13, value=b.priority_level or '')

        for col in range(1, len(header) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"summary_{datetime.date.today().isoformat()}.xlsx"
        return (bio.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"'
        })

    @app.route('/reports/summary/download_pdf')
    @login_required
    def reports_summary_pdf():
        query = build_bid_query_from_args(request.args)
        bids = query.order_by(Bid.date_monitored.desc().nullslast(), Bid.closing_date.asc().nullslast()).all()
        html = render_template('summary_report.html', bids=bids, filters=request.args, pdf=True, generated_at=datetime.datetime.utcnow())
        # Try pdfkit (wkhtmltopdf), then WeasyPrint, otherwise show printable view with instructions
        if _PDFKIT_AVAILABLE:
            try:
                # Allow overriding wkhtmltopdf path via environment variable
                wkpath = os.environ.get('WKHTMLTOPDF_PATH') or app.config.get('WKHTMLTOPDF_PATH')
                pdf_conf = None
                try:
                    if wkpath:
                        pdf_conf = pdfkit.configuration(wkhtmltopdf=wkpath)
                    else:
                        pdf_conf = None
                except Exception:
                    pdf_conf = None
                if pdf_conf:
                    pdf = pdfkit.from_string(html, False, configuration=pdf_conf)
                else:
                    pdf = pdfkit.from_string(html, False)
                filename = f"summary_{datetime.date.today().isoformat()}.pdf"
                return (pdf, 200, {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': f'attachment; filename="{filename}"'
                })
            except Exception:
                flash('PDF generation via pdfkit failed on the server. Trying alternative...', 'warning')
        if _WEASY_AVAILABLE:
            try:
                pdf = WeasyHTML(string=html).write_pdf()
                filename = f"summary_{datetime.date.today().isoformat()}.pdf"
                return (pdf, 200, {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': f'attachment; filename="{filename}"'
                })
            except Exception:
                flash('PDF generation via WeasyPrint failed on the server. Showing printable view instead.', 'warning')

        flash('PDF generation is not configured on this server. Install `wkhtmltopdf` and enable `pdfkit` or install `WeasyPrint` for server-side PDFs. Use Print to save as PDF from your browser.', 'warning')
        return render_template('summary_report.html', bids=bids, filters=request.args, generated_at=datetime.datetime.utcnow())

    @app.route('/bid-monitoring')
    @login_required
    def bid_monitoring():
        search = request.args.get('search', '', type=str).strip()
        date_added_filter = request.args.get('date_added', '', type=str).strip()
        closing_date_filter = request.args.get('closing_date', '', type=str).strip()
        status_filter = request.args.get('status', '', type=str).strip()
        region_filter = request.args.get('region', '', type=str).strip()
        category_filter = request.args.get('category', '', type=str).strip()
        priority_filter = request.args.get('priority', '', type=str).strip()
        system_type_filter = request.args.get('system_type', '', type=str).strip()
        closing_soon_filter = request.args.get('closing_soon', '', type=str).strip()

        query = Bid.query

        if search:
            like_term = f'%{search}%'
            query = query.filter(
                or_(
                    Bid.reference_number.ilike(like_term),
                    Bid.control_number.ilike(like_term),
                    Bid.project_title.ilike(like_term),
                    Bid.procuring_entity.ilike(like_term),
                )
            )

        canvassing_filter = request.args.get('canvassing_status', '', type=str).strip()

        # apply specific date_added filter (exact date)
        if date_added_filter:
            parsed = parse_date_input(date_added_filter)
            if parsed:
                query = query.filter(Bid.date_added == parsed)

        # apply specific closing date filter (exact date)
        if closing_date_filter:
            parsed = parse_date_input(closing_date_filter)
            if parsed:
                query = query.filter(Bid.closing_date == parsed)
                if not status_filter:
                    query = query.filter(Bid.status != 'Cancelled')

        if status_filter:
            query = query.filter(Bid.status == status_filter)
        if region_filter:
            query = query.filter(Bid.region == region_filter)
        if category_filter:
            query = query.filter(Bid.category == category_filter)
        if priority_filter:
            query = query.filter(Bid.priority_level == priority_filter)
        if system_type_filter:
            query = query.filter(Bid.system_type == system_type_filter)
        if canvassing_filter:
            query = query.filter(Bid.canvassing_status == canvassing_filter)
        if closing_soon_filter == 'true':
            today = datetime.date.today()
            next_week = today + datetime.timedelta(days=7)
            query = query.filter(
                Bid.closing_date.isnot(None),
                Bid.closing_date >= today,
                Bid.closing_date <= next_week,
                Bid.status != 'Cancelled',
            )

        page = request.args.get('page', 1, type=int)
        pagination = query.order_by(Bid.date_monitored.desc().nullslast(), Bid.reference_number.asc()).paginate(
            page=page, per_page=25, error_out=False
        )

        statuses = [row[0] for row in db.session.query(Bid.status).distinct().order_by(Bid.status).all() if row[0]]
        regions = [row[0] for row in db.session.query(Bid.region).distinct().order_by(Bid.region).all() if row[0]]
        db_categories = [row[0] for row in db.session.query(Bid.category).distinct().order_by(Bid.category).all() if row[0]]
        categories = resolve_categories(db_categories)
        priorities = [row[0] for row in db.session.query(Bid.priority_level).distinct().order_by(Bid.priority_level).all() if row[0]]
        canvassing_statuses = [row[0] for row in db.session.query(Bid.canvassing_status).distinct().order_by(Bid.canvassing_status).all() if row[0]]

        system_types = [row[0] for row in db.session.query(Bid.system_type).distinct().order_by(Bid.system_type).all() if row[0]]

        return render_template(
            'bid_monitoring.html',
            bids=pagination.items,
            pagination=pagination,
            filters={
                'search': search,
                'date_added': date_added_filter,
                'closing_date': closing_date_filter,
                'status': status_filter,
                'region': region_filter,
                'category': category_filter,
                'priority': priority_filter,
                'system_type': system_type_filter,
                'canvassing_status': canvassing_filter,
                'closing_soon': closing_soon_filter,
            },
            statuses=statuses,
            regions=regions,
            categories=categories,
            priorities=priorities,
            canvassing_statuses=canvassing_statuses,
            system_types=system_types,
        )

    @app.route('/bid-monitoring/<int:bid_id>')
    @login_required
    def bid_detail(bid_id):
        bid = Bid.query.get_or_404(bid_id)
        return render_template('bid_detail.html', bid=bid)

    return app


if __name__ == '__main__':
    app = create_app()
    with app.app_context():
        db.create_all()
        from sqlalchemy import inspect as _inspect
        inspector = _inspect(db.engine)
        if 'bids' in inspector.get_table_names():
            columns = [column['name'] for column in inspector.get_columns('bids')]
            if 'attachments' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text('ALTER TABLE bids ADD COLUMN attachments VARCHAR(2048)'))
    app.run(host='0.0.0.0', port=5001, debug=True)
