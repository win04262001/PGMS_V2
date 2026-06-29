from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import datetime

from app import create_app
from models import db, Bid

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / 'PhilGEPS_Monitoring_System_v2_Fixed.xlsx'

COLUMN_MAPPING = {
    0: 'date_monitored',
    1: 'reference_number',
    2: 'procuring_entity',
    3: 'project_title',
    4: 'category',
    5: 'region',
    6: 'province',
    7: 'abc_value',
    8: 'closing_date',
    9: 'closing_time',
    10: 'days_remaining',
    11: 'priority_level',
    13: 'status',
    14: 'remarks',
    15: 'date_submitted',
    16: 'bid_amount',
    17: 'result',
    18: 'opportunity_score',
    19: 'abc_flag',
    20: 'recommendation',
}


def normalize_text(value):
    if value is None:
        return None
    return str(value).strip()


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            return dt.date() if hasattr(dt, 'date') else None
        except Exception:
            return None
    text = str(value).strip()
    for parser in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(text, parser).date()
        except ValueError:
            pass
    return None


def parse_time(value):
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value.strftime('%H:%M')
    if isinstance(value, datetime.datetime):
        return value.time().strftime('%H:%M')
    return str(value).strip()


def parse_int(value):
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def parse_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        text = str(value).replace(',', '').strip()
        try:
            return float(text)
        except (TypeError, ValueError):
            return None


def build_row(row_values):
    row_data = {}
    for index, field_name in COLUMN_MAPPING.items():
        value = row_values[index] if index < len(row_values) else None
        if field_name in {'date_monitored', 'closing_date', 'date_submitted'}:
            row_data[field_name] = parse_date(value)
        elif field_name in {'abc_value', 'bid_amount'}:
            row_data[field_name] = parse_float(value)
        elif field_name in {'days_remaining', 'opportunity_score'}:
            row_data[field_name] = parse_int(value)
        elif field_name == 'closing_time':
            row_data[field_name] = parse_time(value)
        else:
            row_data[field_name] = normalize_text(value)
    return row_data


if __name__ == '__main__':
    wb = load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Bid Monitoring']
    app = create_app()

    with app.app_context():
        db.create_all()

        imported = 0
        updated = 0

        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not any(cell is not None for cell in row[:22]):
                continue

            row_data = build_row(row)
            reference = row_data.get('reference_number')
            if not reference:
                continue

            bid = Bid.query.filter_by(reference_number=reference).first()
            if bid is None:
                bid = Bid(**row_data)
                db.session.add(bid)
                imported += 1
            else:
                for key, value in row_data.items():
                    setattr(bid, key, value)
                updated += 1

        db.session.commit()
        print(f'Imported: {imported}, Updated: {updated}')
